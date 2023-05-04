import openpyxl
import pandas as pd
from selenium import webdriver
from time import sleep
import xlwings as xw
import openpyxl

username = "1234"
password = "1234"
mots = "1234"

url = "https:/*"
driver =webdriver.Chrome("C:\chromedriver.exe")

driver.get(url)

driver.find_element_by_id("GloATTUID").send_keys(username)
driver.find_element_by_id("GloPassword").send_keys(password)
driver.find_element_by_id("GloPasswordSubmit").click()
driver.find_element_by_id("successButtonId").click()

print("Logged in Successfully")
sleep(5)

wb = openpyxl.load_workbook('Incidents.xlsx')
sh1 = wb['Sheet1']
row = sh1.max_row
print("total URLs : " , row)
column = sh1.max_column

driver.get("https://cana.web.*")
sleep(4)
driver.find_element_by_id("successButtonId").click()

try:
 for i in range(1,row+1):
    for j in range(1,column+1):
            currentdata = sh1.cell(i, j).value
            print(currentdata)

            driver.get("https://onetool.it.att.com/om/ActiveIncidentTemplate.cfm?PRNUM=" + str(currentdata) + "&Incident=1")


            titles = driver.find_elements_by_xpath('')
            for title in titles :
                    print(title.text)




            #degradation_min = driver.find_element_by_xpath("").text

            sleep(5)


except Exception as e:
    print(e)

print("Success")
