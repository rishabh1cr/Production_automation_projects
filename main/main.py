import openpyxl
import xlwings as xw
import openpyxl
import requests
from selenium import webdriver
from time import sleep
import  webbrowser
from openpyxl.styles import PatternFill

import os
import win32com.client as client
from PIL import ImageGrab

from PIL import Image, ImageChops
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
import time

options =webdriver.ChromeOptions()
options.headless= True
driver =webdriver.Chrome("C:\Program Files\chromedriver.exe", options=options)

#limit =['1','2','3']
wb = openpyxl.load_workbook('C:\\Users\\rishabh\\Desktop\\CJ Automation\\Customer_Journey_Automation.xlsx')
sh1 = wb['DirectURLHit']

row = sh1.max_row
print("Total number of URLs to be checked : " , row)
column = sh1.max_column

for i in range(1,row+1):
    for j in range(1,column+1):
        currentdata=sh1.cell(i,j).value
        driver.execute_script("window.open()")
        driver.switch_to.window(driver.window_handles[i])
        print("Current window : " , i)
        driver.get(currentdata)
        sleep(8)
        S = lambda X: driver.execute_script('return document.body.parentNode.scroll' + X)
        driver.set_window_size(S('Width'), S('Height'))
        driver.find_element_by_tag_name('body').screenshot('C:\\Users\\rishabh\\Desktop\\Photos\\DirectURLs\\TestData\\f_' + str(i) + '.png')
        print("Current URl number under validation : ", i)


    img1 = Image.open('C:\\Users\\rishabh\\Desktop\\Photos\\DirectURLs\\TestData\\f_' + str(i) + '.png')
    img2 = Image.open('C:\\Users\\rishabh\\Desktop\\Photos\\DirectURLs\\ActualData\\file_' + str(i) + '.png')

    diff = ImageChops.difference(img1, img2)
    if diff.getbbox():
        diff.show()
        r = sh1.cell(i, j + 1, value="Fail")
        r.fill = PatternFill("solid", fgColor="F50707")
        print("The pictures does not match")

    else :
        g = sh1.cell(i, j + 1, value="Success")
        g.fill = PatternFill("solid", fgColor="71ff33")
        print("Validated Successfully ")
    wb.save("Reports.xlsx")

wb.save("Reports.xlsx")
sleep(5)
openpyxl.load_workbook("Reports.xlsx")


# email code
workbook_path = os.getcwd() + '\\Reports.xlsx'
excel = client.Dispatch('Excel.Application')
wb = excel.Workbooks.Open(workbook_path)

sheet = wb.Sheets['DirectURLHit']
excel.Visible = 1
copyrange = sheet.Range('A1:B23')
copyrange.CopyPicture(Appearance=1, Format=2)
ImageGrab.grabclipboard().save('DirectURLImage.png')

#Create mail
image_path = os.getcwd() + '\\DirectURLImage.png'
html_body = """
<div>
Please refer below . <br><br>
</div>
<div>
<img src={}></img>
</div>
"""

outlook = client.Dispatch('Outlook.Application')
message = outlook.CreateItem(0)
message.To = "rishabh@*.com"
message.Subject = "Please review"
message.HTMLBody = html_body.format(image_path)
message.Display()
message.Save()
